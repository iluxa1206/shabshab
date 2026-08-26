"""Снятие manual_locked со строк реестра, залоченных АВТОМАТИЧЕСКИ (round-trip xlsx).

ЗАЧЕМ. Импорт xlsx в Справочнике зовёт set_manual(lock=True) на КАЖДУЮ строку файла,
даже если пользователь ничего в ней не менял (типичный сценарий: выгрузил каталог,
загрузил обратно). manual_locked=1 запрещает upsert обновлять _MANUAL_FIELDS —
base, margin_bps, maturity_date, issue_date, face_value, day_count и т.д. В итоге
sync перестаёт подтягивать свежие maturity/margin/номинал из MOEX/Cbonds, и данные
тихо устаревают. Ср. scripts/unfreeze_fixing_spec.py — там та же причина, но про
поля спеки фиксинга.

ЧТО ДЕЛАЕТ. Сверяет ключевые поля залоченной строки с Cbonds-выгрузкой
(bondsearch_*.xlsx, тот же источник, что и у sync). Совпадает → строка НЕ ручная,
снимаем lock. Расходится → возможно ручная правка, ОСТАВЛЯЕМ залоченной и печатаем.

ЗАПУСК (в контейнере): dry-run по умолчанию, APPLY=1 — запись + бэкап.
    docker compose -f docker-compose.prod.yml exec floaters \
        python scripts/unlock_auto_imported.py
    docker compose -f docker-compose.prod.yml exec -e APPLY=1 floaters \
        python scripts/unlock_auto_imported.py
"""
from __future__ import annotations

import glob
import os
import shutil
import sqlite3
import sys
from datetime import datetime, timezone

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from services.paths import CACHE_DIR  # noqa: E402

DB_PATH = os.environ.get("INSTRUMENTS_DB") or os.path.join(
    os.path.dirname(CACHE_DIR), "instruments.db")


def _load_bondsearch() -> dict:
    """{isin: {base, margin_bps, maturity}} из свежайшего bondsearch_*.xlsx."""
    files = glob.glob(os.path.join(_ROOT, "bondsearch_*.xlsx"))
    if not files:
        return {}
    path = max(files, key=os.path.getmtime)
    print(f"источник сверки: {os.path.basename(path)}")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)
    out = {}
    for r in it:
        if not r or not r[1]:
            continue
        base_raw = str(r[32] or "").upper()
        base = ("RUONIA" if "RUONIA" in base_raw and "ИНДЕКС" not in base_raw
                else "KEYRATE" if "КЛЮЧЕВАЯ" in base_raw else None)
        try:
            margin = int(round(float(r[33]) * 100))
        except (TypeError, ValueError):
            margin = None
        out[str(r[1]).strip()] = {"base": base, "margin_bps": margin}
    return out


def main() -> int:
    apply = os.environ.get("APPLY") == "1"
    if not os.path.exists(DB_PATH):
        print(f"БД не найдена: {DB_PATH}")
        return 1
    bs = _load_bondsearch()
    if not bs:
        print("bondsearch_*.xlsx не найден — сверять не с чем, выходим")
        return 1

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """SELECT isin, short_name, base, margin_bps, coupon_mode, fixing_lag,
                  cap_pct, floor_pct
           FROM instruments WHERE manual_locked=1"""
    ).fetchall()

    unlock, keep, unknown, spec = [], [], [], []
    for r in rows:
        src = bs.get(r["isin"])
        if not src or src["base"] is None:
            unknown.append(r)          # нет в выгрузке — сверить не с чем
            continue
        # СПЕКА ФИКСИНГА — ОТДЕЛЬНЫЙ СЛУЧАЙ. Сверяем мы только base/margin_bps,
        # а lock защищает весь _MANUAL_FIELDS, включая coupon_mode/fixing_lag.
        # Разлочив такую строку, мы открываем её для apply_authoritative:
        # corpbonds ПЕРЕЗАПИШЕТ режим купона (он игнорирует locked, но пишет
        # поверх разлоченного). Спека могла быть выставлена руками или
        # заморожена старым парсером — решать это должен человек, а не сверка
        # по двум полям. Разморозку спеки делает scripts/unfreeze_fixing_spec.py.
        # cap_pct/floor_pct — тоже ручной слой: apply_authoritative пишет любые
        # переданные поля. var_type СЮДА НЕ ВХОДИТ — он авто-проставлен почти
        # всем строкам и признаком ручной правки не является.
        if not os.getenv("UNLOCK_WITH_SPEC") and (
                r["coupon_mode"] is not None or r["cap_pct"] is not None
                or r["floor_pct"] is not None):
            spec.append((r, src))
            continue
        same_base = (r["base"] == src["base"])
        same_margin = (src["margin_bps"] is None or r["margin_bps"] == src["margin_bps"])
        (unlock if (same_base and same_margin) else keep).append((r, src))

    print(f"залочено строк              : {len(rows)}")
    print(f"  совпадает с Cbonds → СНЯТЬ: {len(unlock)}")
    print(f"  расходится → ОСТАВИТЬ     : {len(keep)}")
    print(f"  нет в выгрузке → ОСТАВИТЬ : {len(unknown)}")
    print(f"  своя спека → ОСТАВИТЬ     : {len(spec)}"
          f"{'  (UNLOCK_WITH_SPEC=1 чтобы снять)' if spec else ''}")
    for r, _src in spec[:25]:
        print(f"    СО СПЕКОЙ {r['isin']} {str(r['short_name'])[:18]:20} "
              f"{r['coupon_mode']}/{r['fixing_lag']}")
    for r, src in keep[:25]:
        print(f"    ОСТАВЛЕНА {r['isin']} {str(r['short_name'])[:18]:20} "
              f"БД={r['base']}/{r['margin_bps']}  Cbonds={src['base']}/{src['margin_bps']}")

    if not apply:
        print("\n[DRY-RUN] записи не было. APPLY=1 — применить.")
        return 0

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup = f"{DB_PATH}.bak-unlock-{stamp}"
    shutil.copy2(DB_PATH, backup)
    print(f"\nбэкап: {backup} ({os.path.getsize(backup)} байт)")

    conn.executemany("UPDATE instruments SET manual_locked=0 WHERE isin=?",
                     [(r["isin"],) for r, _ in unlock])
    conn.commit()
    left = conn.execute("SELECT count(*) n FROM instruments WHERE manual_locked=1").fetchone()["n"]
    print(f"разлочено: {len(unlock)}; осталось залочено: {left}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
