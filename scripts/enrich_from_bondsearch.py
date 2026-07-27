#!/usr/bin/env python3
"""Проставить параметры флоатеров из bondsearch-выгрузки → xlsx-шаблон СПРАВОЧНИКа.

Одноразовый bulk-проход: для каждого РУБ КС/RUONIA-флоатера берём структурные
поля (load_cbonds: база/маржа/погашение/эмиссия/номинал/период/day_count/var_type/
формула) + прогоняем текст формулы теми же парсерами, что кнопка «Разобрать»
(_parse_formula + parse_prospectus_formula) → coupon_mode / fixing_lag / cap / floor.

Выход — instruments_review.xlsx строго в схеме импорта (_XLSX_COLS). Проверяешь
в Excel и заливаешь через СПРАВОЧНИК → «Импорт xlsx» (строки лочатся, sync не
затрёт). Прод-БД не трогается.

Кэп/флор в bondsearch почти всегда обрезаны экспортом (формула ~90–250 симв) —
их полноценно даст только проспект; тут выходят единицы.
"""
import io
import sys
import openpyxl

sys.path.insert(0, ".")
from services import ref_data
from services.enrich_corpbonds import _parse_formula
try:
    from services.coupon_calib import parse_prospectus_formula
except Exception:
    parse_prospectus_formula = lambda _t: {}

# Схема импорта СПРАВОЧНИКа (api/routes/instruments._XLSX_COLS) — заголовки 1:1.
_XLSX_COLS = ("isin", "short_name", "base", "margin_bps", "maturity_date",
              "issue_date", "coupon_period_days", "coupons_per_year", "day_count",
              "face_value", "var_type", "fixing_lag", "fixing_lag_unit", "coupon_mode",
              "cap_pct", "floor_pct", "coupon_text")

BONDSEARCH = sys.argv[1] if len(sys.argv) > 1 else None  # None → свежайший
OUT = sys.argv[2] if len(sys.argv) > 2 else "instruments_review.xlsx"


def _num(v, cast):
    try:
        return cast(float(v))
    except (TypeError, ValueError):
        return None


def parse_formula(txt):
    """Разбор формулы → {coupon_mode, fixing_lag, fixing_lag_unit, cap_pct, floor_pct}.

    mode+lag+unit берём ИЗ ОДНОГО источника — проспект-парсера (авторитет по
    фиксингу): его ключи 'mode'/'lag'/'lag_unit'. Смешивать mode из corpbonds с
    lag из проспекта нельзя — рассинхрон (point с lag от average и наоборот).
    corpbonds-парсер держим только ради флага exotic. Enum фильтруем под
    валидацию импорта (coupon_mode∈point|average, fixing_lag_unit∈cal|work)."""
    out = {}
    try:
        ps = parse_prospectus_formula(txt) or {}
    except Exception:
        ps = {}
    if ps.get("mode") in ("point", "average"):
        out["coupon_mode"] = ps["mode"]
    if ps.get("lag") is not None:
        out["fixing_lag"] = int(ps["lag"])
    if ps.get("lag_unit") in ("cal", "work"):
        out["fixing_lag_unit"] = ps["lag_unit"]
    if ps.get("cap_pct") is not None:
        out["cap_pct"] = float(ps["cap_pct"])
    if ps.get("floor_pct") is not None:
        out["floor_pct"] = float(ps["floor_pct"])
    exotic = (_parse_formula(txt) or {}).get("exotic")
    return out, exotic


def _full_formulas(path):
    """{isin: полный текст формулы} из сырой колонки «Купон». load_cbonds режет
    coupon_text до 300 симв (лимит поля реестра); для РАЗБОРА нужен полный текст
    (окно усреднения/кэп часто в хвосте)."""
    if not path:
        path = ref_data._latest_cbonds_file()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    ii, ic = hdr.index("ISIN"), hdr.index("Купон")
    out = {}
    for row in it:
        if row[ii] and row[ic]:
            out[str(row[ii]).strip().upper()] = str(row[ic])
    return out


def main():
    cb = ref_data.load_cbonds(BONDSEARCH) if BONDSEARCH else ref_data.load_cbonds()
    fl = {k: v for k, v in cb.items() if v.get("base") in ("KEYRATE", "RUONIA")}
    full = _full_formulas(BONDSEARCH)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "instruments"
    ws.append(list(_XLSX_COLS))

    stat = {"rows": 0, "mode": 0, "lag": 0, "cap": 0, "floor": 0, "exotic": 0}
    for isin in sorted(fl):
        v = fl[isin]
        freq = _num(v.get("freq"), float)
        # разбор по ПОЛНОЙ формуле; coupon_text (≤300) — только для поля реестра
        pf, exotic = parse_formula(full.get(isin) or v.get("coupon_text") or "")
        row = {
            "isin": isin,
            "short_name": v.get("name"),
            "base": v.get("base"),
            "margin_bps": _num(v.get("margin_bps"), int),
            "maturity_date": v.get("maturity_date"),
            "issue_date": v.get("issue_date"),
            "coupons_per_year": int(freq) if freq else None,
            "coupon_period_days": round(365 / freq) if freq else None,
            "day_count": v.get("day_count"),
            "face_value": _num(v.get("face_value"), float),
            "var_type": v.get("var_type"),
            "coupon_text": v.get("coupon_text"),
            **pf,
        }
        ws.append([row.get(c) for c in _XLSX_COLS])
        stat["rows"] += 1
        for k in ("coupon_mode", "fixing_lag", "cap_pct", "floor_pct"):
            if row.get(k) is not None:
                stat[{"coupon_mode": "mode", "fixing_lag": "lag",
                      "cap_pct": "cap", "floor_pct": "floor"}[k]] += 1
        if exotic:
            stat["exotic"] += 1

    wb.save(OUT)
    print(f"Записано {stat['rows']} флоатеров → {OUT}")
    print(f"  coupon_mode: {stat['mode']}  fixing_lag: {stat['lag']}  "
          f"cap: {stat['cap']}  floor: {stat['floor']}  exotic(⚠): {stat['exotic']}")


if __name__ == "__main__":
    main()
