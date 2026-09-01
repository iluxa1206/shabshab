"""Admin-CRUD реестра инструментов (services.instruments_registry).

Ручной ввод/правка расчётных параметров новых бумаг + список «на ревью».
Всё под require_admin. Расчётные поля тут — то, что нужно нашему прайсингу без НРД.
"""
from __future__ import annotations

import io
import re
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from api.routes.auth import require_admin
from services import instruments_registry as reg

router = APIRouter()

# Колонки справочника-шаблона xlsx (round-trip: экспорт → правка → импорт).
# Порядок = колонки страницы СПРАВОЧНИК (фронт Catalog.jsx COLS), ISIN — ключ.
# Read-only колонки (br_*/spec_eff/rating/source) выгружаются для контекста,
# импорт их игнорирует (_XLSX_EDITABLE).
_XLSX_COLS = ("isin", "short_name", "base", "margin_bps", "maturity_date",
              "issue_date", "coupon_period_days", "coupons_per_year", "day_count",
              "face_value", "coupon_mode", "fixing_lag", "fixing_lag_unit",
              "avg_window_days", "compounded", "br_coupon_mode", "br_fixing_lag", "spec_eff",
              "cap_pct", "floor_pct", "var_type", "coupon_text", "margin_schedule",
              "face_index", "rating", "source")
_XLSX_EDITABLE = ("short_name", "base", "margin_bps", "maturity_date",
                  "issue_date", "coupon_period_days", "coupons_per_year", "day_count",
                  "face_value", "coupon_mode", "fixing_lag", "fixing_lag_unit",
                  "avg_window_days", "compounded", "cap_pct", "floor_pct", "var_type",
                  "coupon_text", "margin_schedule", "face_index")
_XLSX_INT = {"margin_bps", "coupon_period_days", "coupons_per_year", "fixing_lag",
             "avg_window_days", "compounded"}
_XLSX_FLOAT = {"face_value", "cap_pct", "floor_pct"}

# Русские/внешние заголовки, принимаемые импортом наравне с техническими именами
# полей. Ключ — поле реестра, значения — как колонку могут назвать в чужом файле
# (заголовки страницы СПРАВОЧНИК, выгрузка bondsearch). Заголовок нормализуется
# (регистр/пунктуация/пробелы/ё), поэтому «Маржа, бп» = «маржа бп».
# Чтобы принять ещё один формат — допиши сюда алиас, парсер не трогается.
_XLSX_ALIASES = {
    "isin": ["ISIN", "ISIN код"],
    "short_name": ["Название", "Бумага", "Наименование"],
    "base": ["База", "Базовая ставка", "Базовый индекс (для FRN)"],
    # ТОЛЬКО «бп»: в bondsearch колонка «Маржа» — ПРОЦЕНТЫ (2.5), и её алиас
    # тихо записал бы 2 bps. Такой файл заводится через
    # scripts/enrich_from_bondsearch.py, который переводит единицы.
    "margin_bps": ["Маржа, бп", "Спред, бп"],
    "maturity_date": ["Погашение", "Дата погашения"],
    "issue_date": ["Эмиссия", "Начало обращения", "Дата размещения",
                   "Начало начисления купонов"],
    "coupon_period_days": ["Период, дн"],
    "coupons_per_year": ["Куп/год", "Периодичность купона"],
    "day_count": ["Метод расчета НКД", "Базис НКД"],
    "face_value": ["Номинал", "Мин. торг. лот / Номинал"],
    "coupon_mode": ["Режим (БД)", "Режим"],
    "fixing_lag": ["Лаг (БД)", "Лаг"],
    "fixing_lag_unit": ["Ед. лага"],
    "avg_window_days": ["Окно, дн", "Окно"],
    "compounded": ["Капит."],
    "cap_pct": ["Кэп %", "Кэп"],
    "floor_pct": ["Флор %", "Флор"],
    "var_type": ["Тип ставки", "Тип переменной ставки купона"],
    "coupon_text": ["Формула", "Текст формулы"],
    "margin_schedule": ["Лесенка маржи"],
    "face_index": ["Индекс номинала"],
}


def _norm_hdr(v) -> str:
    """Заголовок колонки → канон: нижний регистр, ё→е, без пунктуации."""
    t = str(v or "").lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", " ", t).strip()


def _header_index(header) -> dict:
    """{поле реестра: индекс колонки} по заголовкам файла.

    Сначала техническое имя поля (шаблон экспорта), затем алиасы из
    _XLSX_ALIASES — так один и тот же импорт съедает и наш round-trip, и чужую
    выгрузку с русскими шапками."""
    norm = {}
    for i, h in enumerate(header):
        k = _norm_hdr(h)
        if k and k not in norm:
            norm[k] = i
    idx = {}
    for field in ("isin",) + _XLSX_EDITABLE:
        cands = [field] + _XLSX_ALIASES.get(field, [])
        for nm in cands:
            hit = norm.get(_norm_hdr(nm))
            if hit is not None:
                idx[field] = hit
                break
    return idx


_ISIN_RE = re.compile(r"[A-Z]{2}[A-Z0-9]{9}[0-9]")


def _require_isin(isin: str) -> str:
    isin = (isin or "").strip().upper()
    if not _ISIN_RE.fullmatch(isin):
        raise HTTPException(status_code=422, detail="Некорректный ISIN")
    return isin


class InstrumentParams(BaseModel):
    """Ручные параметры бумаги. Все опциональны — правим только заданные поля."""
    base: Optional[str] = Field(None, description="KEYRATE | RUONIA | FIXED")
    margin_bps: Optional[int] = None
    maturity_date: Optional[str] = Field(None, description="ISO YYYY-MM-DD")
    issue_date: Optional[str] = None
    coupon_period_days: Optional[int] = Field(None, ge=1, le=1830)
    coupons_per_year: Optional[int] = Field(None, ge=1, le=365)
    day_count: Optional[str] = None
    face_value: Optional[float] = Field(None, gt=0)
    # лаг до 400: конвенции вида «среднее предыдущего периода» кодируются большим
    # лагом (полугодовой купон → лаг ~182-190, см. bondresearch)
    fixing_lag: Optional[int] = Field(None, ge=0, le=400)
    fixing_lag_unit: Optional[str] = Field(None, description="cal | work")
    coupon_mode: Optional[str] = Field(
        None, description="average | month_start (легаси point/avg_prev принимаются и конвертятся: "
                          "point → average+окно 1, avg_prev → average+окно=период)")
    avg_window_days: Optional[int] = Field(
        None, ge=1, le=400,
        description="окно усреднения базы, дней: 1=точечный фиксинг, пусто=длина купонного периода")
    compounded: Optional[int] = Field(
        None, ge=0, le=1,
        description="1 — индекс капитализируется внутри периода (Index_end/Index_start: ВЭБ.РФ, ОФЗ-ПК 29026+)")
    short_name: Optional[str] = Field(None, max_length=128)
    var_type: Optional[str] = None
    cap_pct: Optional[float] = Field(None, ge=0, le=100, description="потолок ставки, % год.")
    floor_pct: Optional[float] = Field(None, ge=0, le=100, description="пол ставки, % год.")
    coupon_text: Optional[str] = Field(None, description="текст формулы купона")
    margin_schedule: Optional[str] = Field(
        None, max_length=512,
        description="лесенка маржи по номерам купонов: «7-20=400; 21-24=550» (bps) "
                    "или JSON [{\"from\":7,\"to\":20,\"bps\":400}]. Купоны вне "
                    "диапазонов = не плавающие (скаляр margin_bps, бэктест их не судит)")
    face_index: Optional[str] = Field(
        None, max_length=16,
        description="ЛИНКЕР: база индексации НОМИНАЛА — 'RUONIA' или пусто. У такой "
                    "бумаги ставка купона фиксирована (её и держит margin_bps), а по "
                    "индексу растёт номинал. Детект автоматический "
                    "(services.linker), это поле — ручной оверрайд")


_ISO_RE = re.compile(r"\d{4}-\d{2}-\d{2}")


def _norm_margin_schedule(v):
    """Ручная лесенка маржи → канонический «7-20=400; 21-24=550» (или "" для
    очистки). Кидает ValueError с человеческим текстом — в БД не должна попасть
    строка, которую прайсинг потом молча не разберёт."""
    from services.coupon_calib import parse_margin_schedule_field
    steps = parse_margin_schedule_field(v)
    if not steps:
        return ""
    return "; ".join(f"{st['from']}-{st['to']}={st['bps']}" for st in steps)


def _offers_no_spec() -> list:
    """Прайсуемые флоатеры с БУДУЩЕЙ офертой и без спеки поведения на ней
    (var_type пуст, cut_at_offer не задан): поток моделируется до погашения по
    дефолту «не резать» — возможно неверно, если эмитент пересматривает купон.
    Ручной механизм уже есть (var_type/cut_at_offer в Справочнике) — здесь только
    видимость кандидатов. Оферты из day-кэша расписаний, без сети."""
    from datetime import date as _date
    from services.market_data import MarketDataService as M
    from services import ref_data
    today = _date.today().isoformat()
    out = []
    for r in reg.universe_rows(only_priceable=True):
        full = M.cached_schedule(r["isin"]) or {}
        future = [o.get("date") for o in full.get("offers", [])
                  if (o.get("date") or "") > today]
        if not future:
            continue
        p = ref_data.params(r["isin"])
        if p.get("var_type") or p.get("cut_at_offer") is not None:
            continue
        out.append({"isin": r["isin"], "short_name": r["name"],
                    "offer_date": min(future)})
    out.sort(key=lambda x: x["offer_date"])
    return out


@router.get("/unreviewed", tags=["Instruments"])
async def unreviewed(_admin: dict = Depends(require_admin)):
    """Новые на ревью + непрайсуемые (нет base/margin/maturity) + suspect (маржа
    расходится с фактом КС/RUONIA) + offer_reset (ставка менялась на прошлой
    оферте, впереди ещё одна, поток к ней не режется) + экзотика (с формулой —
    ловить ложные) + оферты без спеки поведения + счётчики."""
    return {"items": reg.list_unreviewed(),
            "incomplete": reg.list_incomplete(),
            "suspect": reg.list_suspect(),
            "offer_reset": reg.list_offer_reset(),
            "exotic": reg.list_exotic(),
            "offers_no_spec": _offers_no_spec(),
            "count": reg.count()}


@router.get("/new-issues", tags=["Instruments"])
async def new_issues(days: int = Query(0, ge=0, le=365),
                     _admin: dict = Depends(require_admin)):
    """Свежие выпуски (моложе NEW_ISSUE_DAYS) — очередь ручной проверки параметров.
    n — счётчик неподтверждённых, он же значок на кнопке «Справочник»."""
    rows = reg.list_new_issues(days or reg.NEW_ISSUE_DAYS)
    return {"items": rows, "days": days or reg.NEW_ISSUE_DAYS,
            "n": sum(1 for r in rows if not r["reviewed"]),
            "blind": sum(1 for r in rows if not r["priceable"])}


@router.get("/catalog", tags=["Instruments"])
async def catalog(only_active: bool = True, floaters_only: bool = False,
                  _admin: dict = Depends(require_admin)):
    """Полный справочник бумаг со всеми параметрами (спарсенные + пропуски).
    Непрайсуемые вперёд. Для страницы «Справочник» (ручная правка/импорт).
    cbonds_id прикрепляем из bondsearch (прямая ссылка на страницу выпуска)."""
    items = reg.list_catalog(only_active=only_active, floaters_only=floaters_only)
    from services.ref_data import load_cbonds
    cb = load_cbonds()
    for it in items:
        it["cbonds_id"] = (cb.get(it["isin"]) or {}).get("cbonds_id")
    # свежие выпуски: подсветка строк + фильтр «новые» — параметры такой бумаги
    # источники доливают неделями, глаз админа обязателен
    new_rows = reg.list_new_issues()
    # offers_no_spec — прайсуемые с будущей офертой без спеки поведения
    # (var_type/cut_at_offer): горизонт по дефолту «до погашения», админ должен
    # видеть кандидатов
    return {"items": items, "count": reg.count(),
            "new_issues": [r["isin"] for r in new_rows if not r["reviewed"]],
            "new_issue_days": reg.NEW_ISSUE_DAYS,
            "offers_no_spec": _offers_no_spec(),
            "spec_mismatch": [r["isin"] for r in reg.list_spec_mismatch()],
            # тип купона разошёлся с внешним источником (smart-lab): наш вывод о
            # базе проверен НЕ нашими данными, разбирать вручную
            "sl_mismatch": reg.list_sl_mismatch()}


@router.post("/{isin}/recheck-spec", tags=["Instruments"])
async def recheck_spec(isin: str = Path(...), _admin: dict = Depends(require_admin)):
    """Пересчитать бэктест спеки бумаги прямо сейчас (после правки лага/окна):
    прошлые купоны пересчитываются нашей спекой и сверяются с фактом выплат."""
    from datetime import date as _date
    from services import spec_backtest
    isin = _require_isin(isin)
    row = reg.get(isin)
    if row is None:
        raise HTTPException(status_code=404, detail="Нет в реестре")
    r = await spec_backtest._one(isin, row, _date.today())
    if r is None:
        raise HTTPException(status_code=422, detail="Бэктест неприменим (база/расписание)")
    reg.set_spec_backtest(isin, r["err"], r["verdict"], r["n"])
    return {"ok": True, "verdict": r["verdict"], "err_pp": r["err"], "n_coupons": r["n"]}


@router.get("/catalog/export", tags=["Instruments"])
async def catalog_export(only_active: bool = True, floaters_only: bool = False,
                         _admin: dict = Depends(require_admin)):
    """Выгрузка справочника в xlsx — он же ШАБЛОН импорта (правь значения и залей
    обратно). Колонки = редактируемые вручную параметры, ключ ISIN."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "instruments"
    ws.append(list(_XLSX_COLS))
    for r in reg.list_catalog(only_active=only_active, floaters_only=floaters_only):
        ws.append([r.get(c) for c in _XLSX_COLS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=instruments_catalog.xlsx"})


def _coerce(field: str, val):
    """Ячейка xlsx → типизированное значение поля (None — пусто, не трогаем)."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if field in _XLSX_INT:
        return int(float(val))
    if field in _XLSX_FLOAT:
        return float(val)
    return str(val).strip()


def _validate_ranges(params: dict) -> None:
    """Диапазоны полей для xlsx-пути. Бросает ValueError с внятным текстом.

    Ручной POST гоняет параметры через InstrumentParams, а импорт шёл мимо неё:
    ячейка «2,5» (проценты вместо базисных пунктов) писалась как 2 bps и лочилась,
    а maturity_date в прошлом заставляла следующий синк деактивировать бумагу
    навсегда. Границы берём из самой модели, чтобы источник истины был один."""
    from datetime import date as _date
    from pydantic import ValidationError
    try:
        InstrumentParams(**params)
    except ValidationError as e:
        bad = "; ".join(f"{'.'.join(str(x) for x in err['loc'])}: {err['msg']}"
                        for err in e.errors())
        raise ValueError(bad)
    # Модель диапазон дат не знает: погашение в прошлом — это retire_matured
    # на следующем синке, то есть тихое исчезновение бумаги из универса.
    mat = params.get("maturity_date")
    if mat and mat <= _date.today().isoformat():
        raise ValueError(f"maturity_date {mat} не в будущем — бумага будет "
                         "деактивирована следующим синком")
    # margin_bps в модели не ограничена; ловим типовую ошибку «проценты вместо bps»
    mb = params.get("margin_bps")
    if mb is not None and not (-5000 <= int(mb) <= 20000):
        raise ValueError(f"margin_bps {mb} вне −5000..20000 (это базисные пункты, не %)")


def _same_as_registry(cur: dict, field: str, new) -> bool:
    """Значение из xlsx совпадает с тем, что уже лежит в реестре.

    Нужно, чтобы round-trip «выгрузил → поправил одну ячейку → залил» НЕ звал
    set_manual на каждую строку. set_manual — это не «записать значение», а
    «объявить строку ручной»: он ставит manual_locked=1, source='manual' и
    reviewed=1. Экспорт заполняет все поля, поэтому один такой round-trip
    заморозил 538 строк из 1162 — sync перестал обновлять у них номинал,
    погашение и маржу, а очередь ревью вычистилась молча.

    Пустая ячейка сюда не доходит: _coerce отдаёт None, и поле в params не
    попадает. Намеренную очистку мы этим не теряем — импорт её и не умел,
    для этого есть reset_manual."""
    old = cur.get(field)
    if old is None:
        return False                      # заполнение пропуска — это правка
    if field in _XLSX_INT or field in _XLSX_FLOAT:
        try:
            return abs(float(old) - float(new)) <= 1e-9
        except (TypeError, ValueError):
            return False
    return str(old).strip() == str(new).strip()


@router.post("/catalog/import", tags=["Instruments"])
async def catalog_import(file: UploadFile = File(...), _admin: dict = Depends(require_admin)):
    """Импорт параметров из xlsx (шаблон из /catalog/export). Пишем через ручной
    слой (lock — sync не затрёт) ТОЛЬКО РЕАЛЬНО ИЗМЕНЁННЫЕ поля: см. _same_as_registry.
    Возвращает сводку {updated, unchanged, skipped, errors}."""
    import openpyxl
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Не удалось прочитать xlsx")
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise HTTPException(status_code=422, detail="Пустой файл")
    # индексы колонок по заголовку: техническое имя поля ИЛИ русский алиас
    idx = _header_index(header)
    if "isin" not in idx:
        raise HTTPException(status_code=422, detail="Нет колонки ISIN")

    updated, skipped, unchanged, errors = 0, 0, 0, []
    for rn, row in enumerate(rows, start=2):
        raw_isin = row[idx["isin"]] if idx["isin"] < len(row) else None
        isin = (str(raw_isin).strip().upper() if raw_isin else "")
        if not isin:
            continue
        if not _ISIN_RE.fullmatch(isin):
            errors.append(f"строка {rn}: некорректный ISIN «{isin}»")
            continue
        params = {}
        try:
            for field in _XLSX_EDITABLE:
                if field not in idx or idx[field] >= len(row):
                    continue
                v = _coerce(field, row[idx[field]])
                if v is not None:
                    params[field] = v
        except (ValueError, TypeError):
            errors.append(f"строка {rn} ({isin}): числовое поле не распознано")
            continue
        # валидация enum-полей (как в ручном POST)
        # EXOTIC — легальное значение реестра (_FLOAT_BASES_REG), и экспорт его
        # выгружает. Без него round-trip отбрасывал такую строку ЦЕЛИКОМ вместе
        # с правками cap/floor, а errors[:50] прятал хвост списка.
        if params.get("base") and params["base"] not in ("KEYRATE", "RUONIA",
                                                         "FIXED", "EXOTIC"):
            errors.append(f"строка {rn} ({isin}): base ∈ KEYRATE|RUONIA|FIXED|EXOTIC")
            continue
        if params.get("face_index") and params["face_index"] not in ("RUONIA",):
            errors.append(f"строка {rn} ({isin}): face_index ∈ RUONIA либо пусто")
            continue
        if params.get("coupon_mode") and params["coupon_mode"] not in ("point", "average", "avg_prev", "month_start"):
            errors.append(f"строка {rn} ({isin}): coupon_mode ∈ average|month_start "
                          "(point = average+окно1, avg_prev = average+окно=период)")
            continue
        # легаси из старых шаблонов → единая параметризация
        if params.get("coupon_mode") == "point":
            params["coupon_mode"] = "average"
            params.setdefault("avg_window_days", 1)
        elif params.get("coupon_mode") == "avg_prev":
            params["coupon_mode"] = "average"
            if not params.get("avg_window_days"):
                w = params.get("coupon_period_days") or (reg.get(isin) or {}).get("coupon_period_days")
                if w:
                    params["avg_window_days"] = int(w)
                else:
                    errors.append(f"строка {rn} ({isin}): avg_prev без известного периода — задай avg_window_days")
                    continue
        if params.get("fixing_lag_unit") and params["fixing_lag_unit"] not in ("cal", "work"):
            errors.append(f"строка {rn} ({isin}): fixing_lag_unit ∈ cal|work")
            continue
        if "margin_schedule" in params:
            try:
                params["margin_schedule"] = _norm_margin_schedule(params["margin_schedule"])
            except ValueError as e:
                errors.append(f"строка {rn} ({isin}): лесенка маржи — {e}")
                continue
        for f in ("maturity_date", "issue_date"):
            if params.get(f) and not _ISO_RE.fullmatch(params[f]):
                errors.append(f"строка {rn} ({isin}): {f} ждёт YYYY-MM-DD")
                params.pop(f)
        params.pop("isin", None)
        if not params:
            skipped += 1
            continue
        # ТОЛЬКО РАЗЛИЧАЮЩИЕСЯ ПОЛЯ — иначе заморозим строку, где юзер ничего
        # не менял (см. _same_as_registry). Сравниваем ПОСЛЕ нормализации
        # легаси выше, иначе старый шаблон вечно «различается».
        cur = reg.get(isin)
        if cur is not None:
            params = {k: v for k, v in params.items()
                      if not _same_as_registry(cur, k, v)}
            if not params:
                unchanged += 1
                continue
        # ДИАПАЗОНЫ. Ручной POST гоняет их через InstrumentParams, а xlsx-путь
        # шёл мимо: ячейка «2,5» (проценты вместо bps) писалась как 2 bps и
        # лочилась, а maturity_date в прошлом убивала бумагу следующим синком.
        try:
            _validate_ranges(params)
        except ValueError as e:
            errors.append(f"строка {rn} ({isin}): {e}")
            continue
        try:
            reg.set_manual(isin, params, lock=True)
            updated += 1
        except Exception as e:
            errors.append(f"строка {rn} ({isin}): {e}")
    return {"updated": updated, "unchanged": unchanged, "skipped": skipped,
            "errors": errors[:50], "error_count": len(errors)}


class FormulaIn(BaseModel):
    formula: str


@router.post("/parse-formula", tags=["Instruments"])
async def parse_formula(body: FormulaIn, _admin: dict = Depends(require_admin)):
    """Разбор текста формулы купона → {base, margin_bps, coupon_mode, cap_pct,
    floor_pct, fixing_lag, fixing_lag_unit}. Для авто-заполнения полей в СПРАВОЧНИКе.
    Комбинирует парсер corpbonds (база/маржа/режим) + парсер проспекта (лаг/кэп/флор)."""
    from services.enrich_corpbonds import _parse_formula
    out = {}
    f = _parse_formula(body.formula) or {}
    for k in ("base", "margin_bps", "coupon_mode"):
        if f.get(k) is not None:
            out[k] = f[k]
    out["exotic"] = f.get("exotic")   # inverse|capped|None — предупреждение для UI
    try:
        from services.coupon_calib import parse_prospectus_formula
        ps = parse_prospectus_formula(body.formula) or {}
        for k in ("cap_pct", "floor_pct", "fixing_lag_unit"):
            if ps.get(k) is not None:
                out[k if k != "fixing_lag_unit" else "fixing_lag_unit"] = ps[k]
        if ps.get("lag") is not None:
            out["fixing_lag"] = ps["lag"]
        if out.get("coupon_mode") is None and ps.get("mode"):
            out["coupon_mode"] = ps["mode"]
    except Exception:
        pass
    # point убран из модели: точечный фиксинг = average с окном 1 день
    if out.get("coupon_mode") == "point":
        out["coupon_mode"] = "average"
        out.setdefault("avg_window_days", 1)
    return {"parsed": out, "coupon_text": body.formula.strip()}


@router.get("/{isin}", tags=["Instruments"])
async def get_instrument(isin: str = Path(...), _admin: dict = Depends(require_admin)):
    isin = _require_isin(isin)
    row = reg.get(isin)
    if row is None:
        raise HTTPException(status_code=404, detail="Нет в реестре")
    return row


@router.post("/{isin}", tags=["Instruments"])
async def set_instrument(body: InstrumentParams, isin: str = Path(...),
                         _admin: dict = Depends(require_admin)):
    """Ручной ввод/правка параметров (lock — sync их впредь не затрёт) + reviewed."""
    isin = _require_isin(isin)
    params = {k: v for k, v in body.model_dump().items() if v is not None}
    if not params:
        raise HTTPException(status_code=422, detail="Нет полей для сохранения")
    # легаси-вход (старые шаблоны/скрипты) → единая параметризация
    if params.get("coupon_mode") == "point":
        params["coupon_mode"] = "average"
        params.setdefault("avg_window_days", 1)
    elif params.get("coupon_mode") == "avg_prev":
        params["coupon_mode"] = "average"
        if not params.get("avg_window_days"):
            row = reg.get(isin) or {}
            w = params.get("coupon_period_days") or row.get("coupon_period_days")
            if w:
                params["avg_window_days"] = int(w)
            else:
                raise HTTPException(status_code=422,
                                    detail="avg_prev: неизвестен купонный период — задай avg_window_days явно")
    if body.base is not None and body.base not in ("KEYRATE", "RUONIA", "FIXED"):
        raise HTTPException(status_code=422, detail="base ∈ KEYRATE|RUONIA|FIXED")
    # пустая строка — СНЯТИЕ признака линкера (None в модели значит «поле не
    # прислали», им очистить нельзя)
    if body.face_index is not None and body.face_index not in ("RUONIA", ""):
        raise HTTPException(status_code=422, detail="face_index ∈ RUONIA либо пусто")
    for f in ("maturity_date", "issue_date"):
        if params.get(f) and not _ISO_RE.fullmatch(params[f]):
            raise HTTPException(status_code=422, detail=f"{f}: ожидается YYYY-MM-DD")
    if body.fixing_lag_unit is not None and body.fixing_lag_unit not in ("cal", "work"):
        raise HTTPException(status_code=422, detail="fixing_lag_unit ∈ cal|work")
    if body.coupon_mode is not None and body.coupon_mode not in ("point", "average", "avg_prev", "month_start"):
        raise HTTPException(status_code=422, detail="coupon_mode ∈ point|average|avg_prev|month_start")
    if body.margin_schedule is not None:
        try:
            params["margin_schedule"] = _norm_margin_schedule(body.margin_schedule)
        except ValueError as e:
            raise HTTPException(status_code=422, detail=f"лесенка маржи: {e}")
    reg.set_manual(isin, params, lock=True)
    return {"ok": True, "instrument": reg.get(isin)}


@router.post("/{isin}/reset-manual", tags=["Instruments"])
async def reset_manual(isin: str = Path(...), _admin: dict = Depends(require_admin)):
    """Сброс ручной правки: снять manual_locked + обнулить явные поля спеки
    фиксинга (coupon_mode/fixing_lag/fixing_lag_unit/avg_window_days). Спека
    дальше резолвится авто-источниками (bondresearch > парсер > калибратор);
    расчётные поля остаются и обновляются sync'ом."""
    isin = _require_isin(isin)
    removed = reg.reset_manual(isin)
    if removed is None:
        raise HTTPException(status_code=404, detail="Нет в реестре")
    return {"ok": True, "removed": removed, "instrument": reg.get(isin)}


@router.post("/{isin}/reviewed", tags=["Instruments"])
async def mark_reviewed(isin: str = Path(...), _admin: dict = Depends(require_admin)):
    """Пометить бумагу проверенной без правок (параметры из авто-sync устраивают)."""
    isin = _require_isin(isin)
    if reg.get(isin) is None:
        raise HTTPException(status_code=404, detail="Нет в реестре")
    reg.mark_reviewed(isin)
    return {"ok": True}
