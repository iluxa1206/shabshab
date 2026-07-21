"""Живые данные ЦБ РФ: ключевая ставка (полная дневная история + текущая) и
RUONIA (текущая с cbr.ru + историческая база из сида).

Источники:
  КС      — https://www.cbr.ru/hd_base/KeyRate/ (HTML-таблица, UniDbQuery-диапазон;
            даёт всю дневную историю одним запросом → текущая = последняя строка).
  RUONIA  — https://cbr.ru/hd_base/ruonia/ отдаёт только последние ~2 дня; полной
            истории машинного источника нет (MOEX держит лишь фьючерсы). Поэтому
            история = статический сид (ruonia_seed.json, извлечён из 502_504), а
            текущее значение дотягивается живьём и мёржится.

Кэш на диске (cbr_cache.json) на день + фолбэк на кэш/сид при недоступности ЦБ
(тот же паттерн, что rates.py: stale лучше пустого прода)."""
from __future__ import annotations
import os
import json
import glob
import threading
import datetime as _dt
from datetime import date
from typing import List, Tuple, Optional

import requests
from bs4 import BeautifulSoup
import logging

logger = logging.getLogger(__name__)

_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
from services.paths import cache_path as _cache_path
_CACHE = _cache_path("cbr_cache.json")
_RUONIA_SEED = os.path.join(_DIR, "ruonia_seed.json")
_UA = {"User-Agent": "Mozilla/5.0"}
_KS_URL = "https://www.cbr.ru/hd_base/KeyRate/"
_RUONIA_URL = "https://cbr.ru/hd_base/ruonia/"


def _num(s: str) -> Optional[float]:
    try:
        return float(s.replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (ValueError, AttributeError):
        return None


def _ddmmyyyy(s: str) -> Optional[date]:
    try:
        return _dt.datetime.strptime(s.strip(), "%d.%m.%Y").date()
    except (ValueError, AttributeError):
        return None


def _fetch_ks_live() -> List[Tuple[date, float]]:
    """Полная дневная история КС с cbr.ru → [(date, rate_pct)] по возрастанию дат."""
    frm = "01.01.2013"
    to = date.today().strftime("%d.%m.%Y")
    url = f"{_KS_URL}?UniDbQuery.Posted=True&UniDbQuery.From={frm}&UniDbQuery.To={to}"
    resp = requests.get(url, headers=_UA, timeout=20)
    resp.raise_for_status()
    table = BeautifulSoup(resp.text, "html.parser").find("table")
    out: List[Tuple[date, float]] = []
    for tr in table.find_all("tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
        if len(cells) < 2:
            continue
        d, r = _ddmmyyyy(cells[0]), _num(cells[1])
        if d and r is not None:
            out.append((d, r))
    out.sort(key=lambda x: x[0])
    return out


def _fetch_ruonia_current_live() -> List[Tuple[date, float]]:
    """Последние доступные точки RUONIA с cbr.ru (транспонированная таблица, ~2 дня)."""
    resp = requests.get(_RUONIA_URL, headers=_UA, timeout=15)
    resp.raise_for_status()
    table = BeautifulSoup(resp.text, "html.parser").find("table")
    rows = [[td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            for tr in table.find_all("tr")]
    if len(rows) < 2:
        return []
    dates = [_ddmmyyyy(x) for x in rows[0][1:]]
    rates = [_num(x) for x in rows[1][1:]]
    return [(d, r) for d, r in zip(dates, rates) if d and r is not None]


def _load_seed() -> List[Tuple[date, float]]:
    try:
        with open(_RUONIA_SEED, "r", encoding="utf-8") as f:
            data = json.load(f)
        return [(date.fromisoformat(d), float(v)) for d, v in data["points"]]
    except Exception:
        return []


def _load_rc_ruonia() -> List[Tuple[date, float]]:
    """Дневная история RUONIA из выгрузки ЦБ RC_F*.xlsx (cbr.ru/hd_base/ruonia).
    Колонки: DT (дата), ruo (ставка). Авторитетный источник — закрывает разрыв
    между статическим сидом и live-текущей. Берём свежайший RC_F*-файл."""
    import openpyxl
    files = glob.glob(os.path.join(_DIR, "RC_F*.xlsx"))
    if not files:
        return []
    path = max(files, key=os.path.getmtime)
    out: List[Tuple[date, float]] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(h).strip().lower() if h else "" for h in next(it)]
        di = hdr.index("dt") if "dt" in hdr else 0
        ri = hdr.index("ruo") if "ruo" in hdr else 1
        for row in it:
            d, r = row[di], row[ri]
            if isinstance(d, _dt.datetime):
                d = d.date()
            rv = _num(str(r)) if r is not None else None
            if isinstance(d, date) and rv is not None:
                out.append((d, rv))
    except Exception as e:
        logger.warning(f"RC RUONIA load error: {e}")
    return out


def _load_cache() -> dict:
    try:
        with open(_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def _save_cache(ks: List[Tuple[date, float]], ruonia_live: List[Tuple[date, float]]) -> None:
    try:
        from services.paths import atomic_write_json as _awj
        _awj(_CACHE, {"cache_date": date.today().isoformat(),
                       "ks": [[d.isoformat(), v] for d, v in ks],
                       "ruonia_live": [[d.isoformat(), v] for d, v in ruonia_live]})
    except OSError:
        pass


_mem = {"date": None, "ks": None, "ruonia": None}
_refresh_lock = threading.Lock()


def _refresh() -> None:
    """Наполняет память: КС (live→cache), RUONIA (seed + live current).

    Лок: зовётся и из to_thread (warmup), и синхронно из расчётных путей —
    без лока конкурентные вызовы при протухшем кэше дублировали фетч ЦБ и
    гонялись за _mem."""
    today = date.today().isoformat()
    if _mem["date"] == today and _mem["ks"] is not None:
        return
    with _refresh_lock:
        _refresh_locked(today)


def _refresh_locked(today: str) -> None:
    if _mem["date"] == today and _mem["ks"] is not None:   # re-check под локом
        return
    cache = _load_cache()
    ks: List[Tuple[date, float]] = []
    ruonia_live: List[Tuple[date, float]] = []
    try:
        ks = _fetch_ks_live()
    except Exception as e:
        logger.warning(f"WARNING: CBR KeyRate fetch failed ({e}) — фолбэк на кэш")
    try:
        ruonia_live = _fetch_ruonia_current_live()
    except Exception as e:
        logger.warning(f"WARNING: CBR RUONIA fetch failed ({e})")

    if len(ks) < 100 and cache.get("ks"):  # фетч куцый → stale-кэш
        ks = [(date.fromisoformat(d), v) for d, v in cache["ks"]]
        logger.warning("WARNING: CBR KeyRate stale cache fallback")
    if not ruonia_live and cache.get("ruonia_live"):
        ruonia_live = [(date.fromisoformat(d), v) for d, v in cache["ruonia_live"]]

    if ks or ruonia_live:
        _save_cache(ks, ruonia_live)

    # RUONIA = сид (2010→) + выгрузка ЦБ RC_F (2024→, авторитетно) + live current,
    # мёрж по дате с возрастающим приоритетом (live > RC > сид) — разрыв закрыт
    merged = {d: v for d, v in _load_seed()}
    for d, v in _load_rc_ruonia():
        merged[d] = v
    for d, v in ruonia_live:
        merged[d] = v
    ruonia = sorted(merged.items())

    _mem.update({"date": today, "ks": ks, "ruonia": ruonia})


def ks_history() -> List[Tuple[date, float]]:
    _refresh()
    return _mem["ks"] or []


def ruonia_history() -> List[Tuple[date, float]]:
    _refresh()
    return _mem["ruonia"] or []


def current_ks() -> Optional[float]:
    h = ks_history()
    return h[-1][1] if h else None


def ks_rate_at(d: date) -> Optional[float]:
    """Факт КС (decimal, напр. 0.1425) на дату d — последнее значение ЦБ ≤ d.
    Для фиксинга купона KEYRATE-флоатера: ставка купона = КС(дата фиксинга)+маржа."""
    val = None
    for md, r in ks_history():
        if md <= d:
            val = r
        else:
            break
    return (val / 100.0) if val is not None else None


def current_ruonia() -> Optional[float]:
    h = ruonia_history()
    return h[-1][1] if h else None
