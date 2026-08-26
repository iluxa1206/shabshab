"""Справочник параметров облигаций: слой Cbonds-выгрузки + ручные оверрайды.

Разрешение параметров (по убыванию приоритета):
  base, margin, day_count, freq  ← manual > Cbonds-файл > (НРД/MOEX выше по стеку)
  fixing_lag, coupon_mode        ← manual > эмпирический калибратор (coupon_calib)

Cbonds-выгрузка (bondsearch_*.xlsx) даёт базу/маржу/day-count/частоту/тип, но НЕ
лаг фиксинга и не конвенцию point/average — их нет ни у кого, кроме эмиссионных
доков; закрываются калибратором (из истории купонов) или ручным вводом.

Ручной слой (bond_params_manual.json) — редактируемый: новые выпуски + любое
недостающее (лаг, конвенция, маржа для экзотических баз)."""
from __future__ import annotations
import os
import json
import glob
import re
from typing import Dict, Optional

import logging

import openpyxl

logger = logging.getLogger(__name__)

_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_MANUAL_FILE = os.path.join(_DIR, "bond_params_manual.json")

# База Cbonds -> наш тип
_BASE_MAP = {
    "ключевая ставка цб рф": "KEYRATE",
    "ключевая ставка": "KEYRATE",
    "cbr_rate": "KEYRATE",
    "ruonia": "RUONIA",
    "ruonia индекс": "RUONIA",
    "1/2_cbr_rate": "KEYRATE",  # половинная — обрабатывать отдельно, помечаем базу
}


def _cell_iso(v) -> Optional[str]:
    """Ячейка bondsearch (datetime или 'дд.мм.гггг') → ISO YYYY-MM-DD."""
    if v is None or v == "":
        return None
    if hasattr(v, "year") and hasattr(v, "month"):   # datetime/date
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", str(v).strip())
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else None

# Имена колонок в bondsearch-выгрузке (сопоставление по заголовку, не по индексу)
_COL = {
    "isin": ["ISIN"],
    "name": ["Бумага", "Название"],
    "base": ["Базовая ставка", "Базовый индекс (для FRN)"],
    "margin": ["Маржа", "Премия/Дисконт к базовому индексу (для FRN)"],
    "day_count": ["Метод расчета НКД"],
    "freq": ["Периодичность купона", "Купон (раз/год)"],
    "var_type": ["Тип переменной ставки купона"],
    "put": ["Оферта (put)"],
    "coupon_text": ["Купон"],   # текст формулы из проспекта: режим фиксинга + лаг
    "maturity": ["Погашение"],
    "issue": ["Начало начисления купонов", "Начало обращения"],
    "face": ["Мин. торг. лот / Номинал", "Номинал"],
    "cbonds_id": ["Cbonds ID"],   # для прямой ссылки cbonds.ru/bonds/{id}/
}

_cbonds_cache: Optional[Dict[str, dict]] = None
_manual_cache: Optional[Dict[str, dict]] = None


def _hdr_index(headers: list) -> Dict[str, int]:
    idx = {}
    for key, names in _COL.items():
        for nm in names:
            if nm in headers:
                idx[key] = headers.index(nm)
                break
    return idx


def _to_float(v) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _latest_cbonds_file() -> Optional[str]:
    """Свежайший bondsearch_DD_MM_YYYY.xlsx по ДАТЕ в имени (не по строке —
    иначе 20_11_2025 «больше» 10_07_2026). Фолбэк — mtime."""
    import re
    files = glob.glob(os.path.join(_DIR, "bondsearch_*.xlsx"))
    if not files:
        return None

    def key(f):
        m = re.search(r"bondsearch_(\d{2})_(\d{2})_(\d{4})", os.path.basename(f))
        if m:
            dd, mm, yyyy = m.groups()
            return (int(yyyy), int(mm), int(dd))
        return (0, 0, 0)

    dated = [f for f in files if key(f) != (0, 0, 0)]
    if dated:
        return max(dated, key=key)
    return max(files, key=os.path.getmtime)


def load_cbonds(path: Optional[str] = None) -> Dict[str, dict]:
    """{isin: {name, base, margin_bps, day_count, freq, var_type}} из bondsearch-xlsx.
    base∈{KEYRATE,RUONIA,None}; base_raw хранит исходное имя (для экзотики)."""
    global _cbonds_cache
    if _cbonds_cache is not None and path is None:
        return _cbonds_cache
    path = path or _latest_cbonds_file()
    out: Dict[str, dict] = {}
    if not path or not os.path.exists(path):
        _cbonds_cache = out
        return out
    # Файл может быть недочитан (rsync деплоя переписывает его на месте, битая
    # выгрузка, нехватка памяти на воркере): раньше исключение улетало наверх, и
    # coupon_formula отдавал спеку без coupon_mode — прайсинг тихо съезжал на
    # легаси-проекцию. Теперь провал логируется и НЕ кэшируется (следующий вызов
    # пробует снова), а текст формулы подхватывается из реестра (см. params).
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it)]
    except Exception as e:
        logger.error("bondsearch-выгрузка %s не читается: %s — параметры выпусков "
                     "берутся только из реестра/manual", path, e)
        return out
    ix = _hdr_index(headers)
    if "isin" not in ix:
        _cbonds_cache = out
        return out
    for row in it:
        isin = row[ix["isin"]] if ix["isin"] < len(row) else None
        if not isin:
            continue
        base_raw = str(row[ix["base"]]).strip() if ix.get("base") is not None and row[ix["base"]] else None
        base = _BASE_MAP.get((base_raw or "").lower()) if base_raw else None
        margin = _to_float(row[ix["margin"]]) if ix.get("margin") is not None else None
        out[str(isin).strip()] = {
            "name": row[ix["name"]] if ix.get("name") is not None else None,
            "base": base,
            "base_raw": base_raw,
            "margin_bps": round(margin * 100) if margin is not None else None,
            "day_count": row[ix["day_count"]] if ix.get("day_count") is not None else None,
            "freq": _to_float(row[ix["freq"]]) if ix.get("freq") is not None else None,
            "var_type": row[ix["var_type"]] if ix.get("var_type") is not None else None,
            "coupon_text": (str(row[ix["coupon_text"]]).strip()
                            if ix.get("coupon_text") is not None and row[ix["coupon_text"]] else None),
            "maturity_date": _cell_iso(row[ix["maturity"]]) if ix.get("maturity") is not None else None,
            "issue_date": _cell_iso(row[ix["issue"]]) if ix.get("issue") is not None else None,
            "face_value": _to_float(row[ix["face"]]) if ix.get("face") is not None else None,
            "cbonds_id": (int(row[ix["cbonds_id"]]) if ix.get("cbonds_id") is not None
                          and row[ix["cbonds_id"]] not in (None, "") else None),
        }
    if path is None or path == _latest_cbonds_file():
        _cbonds_cache = out
    return out


def load_manual() -> Dict[str, dict]:
    """{isin: {...любые оверрайды: base, margin_bps, fixing_lag, coupon_mode, ...}}."""
    global _manual_cache
    if _manual_cache is not None:
        return _manual_cache
    try:
        with open(_MANUAL_FILE, "r", encoding="utf-8") as f:
            _manual_cache = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        _manual_cache = {}
    return _manual_cache


_reg_ov_cache = {"ts": 0.0, "map": None}


_br_cache = {"ts": 0.0, "map": None}


_txt_cache = {"ts": 0.0, "map": None}


def invalidate_registry_cache() -> None:
    """Сброс кэшей реестровых слоёв — зовётся из instruments_registry при
    ручной правке/импорте, чтобы спека фиксинга применялась сразу, а не через TTL."""
    _reg_ov_cache["map"] = None
    _reg_ov_cache["ts"] = 0.0
    _br_cache["map"] = None
    _br_cache["ts"] = 0.0
    _txt_cache["map"] = None
    _txt_cache["ts"] = 0.0


def _registry_fallback() -> dict:
    """{isin: {coupon_text, var_type, base, margin_bps}} из реестра (все бумаги,
    не только locked) — кэш 30с. САМЫЙ НИЗКИЙ приоритет в params().

    Второй источник параметров, переключающих методику, помимо bondsearch-xlsx.
    Файл xlsx не в git, не в томе данных и приезжает только rsync'ом деплоя: его
    отсутствие раньше тихо обнуляло coupon_mode всему универсу флоатеров и
    признак пересмотра купона у бумаг с офертой (см. pricing_fallback_all)."""
    import time
    now = time.monotonic()
    if _txt_cache["map"] is not None and now - _txt_cache["ts"] < 30:
        return _txt_cache["map"]
    try:
        from services import instruments_registry as reg
        m = reg.pricing_fallback_all()
    except Exception as e:
        logger.warning("реестровый фолбэк параметров недоступен: %s", e)
        m = {}
    _txt_cache["map"] = m
    _txt_cache["ts"] = now
    return m


def _br_specs() -> dict:
    """{isin: {fixing_lag, fixing_lag_unit, coupon_mode}} слоя bondresearch.ru —
    кэш 30с (тот же паттерн, что _registry_overrides)."""
    import time
    now = time.monotonic()
    if _br_cache["map"] is not None and now - _br_cache["ts"] < 30:
        return _br_cache["map"]
    try:
        from services import instruments_registry as reg
        m = reg.br_specs_all()
    except Exception:
        m = {}
    _br_cache["map"] = m
    _br_cache["ts"] = now
    return m


def _registry_overrides() -> dict:
    """{isin: {купонные поля}} из реестра (manual_locked=1) — кэш 30с, чтобы не
    бить SQLite на каждый вызов params() в цикле по юниверсу."""
    import time
    now = time.monotonic()
    if _reg_ov_cache["map"] is not None and now - _reg_ov_cache["ts"] < 30:
        return _reg_ov_cache["map"]
    try:
        from services import instruments_registry as reg
        m = reg.coupon_overrides_all()
    except Exception:
        m = {}
    _reg_ov_cache["map"] = m
    _reg_ov_cache["ts"] = now
    return m


def params(isin: str) -> dict:
    """Слитые параметры выпуска. Приоритет: РЕЕСТР (ручная правка, manual_locked) >
    bond_params_manual.json > Cbonds-справка. Реестровый слой — мост тонкой настройки
    выпуска (формула/кэп/флор/лаг/режим) в прайсинг (см. coupon_formula)."""
    p = dict(load_cbonds().get(isin) or {})
    p.update({k: v for k, v in (load_manual().get(isin) or {}).items() if v is not None})
    p.update(_registry_overrides().get(isin) or {})
    # РЕЕСТРОВЫЙ ФОЛБЭК — последним и только в ПУСТЫЕ поля: свой источник он не
    # затирает (иначе это был бы оверрайд и freeze-trap импорта xlsx). Закрывает
    # дыру, когда bondsearch-выгрузки нет: без coupon_text молча исчезает спека
    # фиксинга, без var_type — обрезка потока по оферте.
    fb = _registry_fallback().get(isin)
    if fb:
        for k, v in fb.items():
            if p.get(k) in (None, ""):
                p[k] = v
    return p


def coupon_formula(isin: str, coupons: list = None, margin_pct: float = None,
                   face: float = None, calc_date=None, amorts: list = None,
                   idx=None) -> dict:
    """Полная спека купона для точной проекции:
        {base, margin_bps, fixing_lag, fixing_lag_unit, coupon_mode}
    Разрешение: base/margin — manual > Cbonds; fixing_lag/coupon_mode —
    manual > ПАРСЕР ПРОСПЕКТА (Cbonds «Купон», авторитетно) > эмпирический
    калибратор (coupon_calib, из истории купонов).
    Возвращает None-поля, если источник не дал (потребитель применяет дефолт)."""
    p = params(isin)
    out = {
        "base": p.get("base"),
        "margin_bps": p.get("margin_bps"),
        "fixing_lag": p.get("fixing_lag"),
        "fixing_lag_unit": p.get("fixing_lag_unit"),
        "coupon_mode": p.get("coupon_mode"),
        "avg_window_days": p.get("avg_window_days"),  # окно усреднения, дней (1=point, NULL=период)
        "compounded": p.get("compounded"),  # капитализация индекса внутри периода (Index_end/Index_start)
        "capped": p.get("capped"),        # есть ли кэп/флор (bool)
        "cap_pct": p.get("cap_pct"),      # потолок ставки купона, % годовых (MIN/«не более»)
        "floor_pct": p.get("floor_pct"),  # пол ставки купона, % годовых (MAX/«не менее»)
        "margin_schedule": None,          # лесенка маржи [{'from','to','bps'}] по № купонов
    }
    # маржа-лесенка из текста формулы: диапазоны купонов со своей надбавкой
    # («S 1-7 = 2.5%, S8-21 = 4.6%»). Скаляр margin_bps остаётся фолбэком для
    # купонов вне распарсенных диапазонов.
    if p.get("coupon_text"):
        try:
            from services.coupon_calib import parse_margin_schedule
            out["margin_schedule"] = parse_margin_schedule(p["coupon_text"])
        except Exception:
            pass
    # 0) ТЕКСТ ПРОСПЕКТА — первым. Это первоисточник: в нём написано, по какому
    #    окну и с каким лагом считается купон. Раньше выше стоял слой
    #    bondresearch.ru, и его «average» затирал распознанный парсером
    #    «avg_prev» у РЖД/РСХБ/Росагро — ошибка бэктеста 0.008 → 0.386 пп
    #    (25.08, регресс приехал с обновлением импорта). Сверка на бэктесте:
    #    из 17 бумаг с расхождением парсер точнее у 1, хуже — ни у одной,
    #    у остальных одинаково.
    if (out["fixing_lag"] is None or out["coupon_mode"] is None
            or out["capped"] is None or out["compounded"] is None) and p.get("coupon_text"):
        try:
            from services.coupon_calib import parse_prospectus_formula
            ps = parse_prospectus_formula(p["coupon_text"])
            if ps:
                if out["coupon_mode"] is None and ps.get("mode") is not None:
                    out["coupon_mode"] = ps["mode"]
                if out["fixing_lag"] is None and ps.get("lag") is not None:
                    out["fixing_lag"] = ps["lag"]
                    out["fixing_lag_unit"] = ps.get("lag_unit", "cal")
                if out["compounded"] is None and ps.get("compounded"):
                    out["compounded"] = 1
                    if out["fixing_lag"] is None and ps.get("lag") is not None:
                        out["fixing_lag"] = ps["lag"]
                if out["capped"] is None:
                    out["capped"] = ps.get("capped", False)
                if out["cap_pct"] is None:
                    out["cap_pct"] = ps.get("cap_pct")
                if out["floor_pct"] is None:
                    out["floor_pct"] = ps.get("floor_pct")
        except Exception:
            pass
    # 2) слой bondresearch.ru (импорт scripts/import_bondresearch_specs.py):
    #    наблюдаемые рынком лаг/метод — ФОЛБЭК для бумаг, где проспекта нет
    #    или парсер его не распознал.
    if out["fixing_lag"] is None or out["coupon_mode"] is None:
        br = _br_specs().get(isin)
        if br:
            if out["fixing_lag"] is None and br.get("fixing_lag") is not None:
                out["fixing_lag"] = br["fixing_lag"]
                out["fixing_lag_unit"] = br.get("fixing_lag_unit", "cal")
            if out["coupon_mode"] is None and br.get("coupon_mode"):
                out["coupon_mode"] = br["coupon_mode"]
                if out["avg_window_days"] is None and br.get("avg_window_days") is not None:
                    out["avg_window_days"] = br["avg_window_days"]
    # 3) фолбэк: калибровка из истории купонов
    if (out["fixing_lag"] is None or out["coupon_mode"] is None) and coupons and calc_date:
        try:
            from services.coupon_calib import calibrate
            m = margin_pct if margin_pct is not None else (out["margin_bps"] or 0) / 100.0
            # режим уже известен (парсер/manual) → фитим только лаг при нём;
            # иначе лаг приезжал из лучшей пары калибратора с ДРУГИМ mode —
            # в прайсинг уходил гибрид «mode парсера + lag чужого фита»
            spec = calibrate(isin, coupons, m, face or 1000.0, calc_date,
                             base=out["base"] or "KEYRATE", amorts=amorts, idx=idx,
                             fixed_mode=out["coupon_mode"])
            if spec:
                if out["fixing_lag"] is None:
                    out["fixing_lag"] = spec["lag"]
                if out["coupon_mode"] is None:
                    out["coupon_mode"] = spec["mode"]
        except Exception:
            pass
    if out["fixing_lag"] is not None and out["fixing_lag_unit"] is None:
        out["fixing_lag_unit"] = "cal"
    # НОРМАЛИЗАЦИЯ (единственная точка резолва): point и avg_prev убраны из
    # модели — всё выражается average + avg_window_days:
    #   point    ≡ average + окно 1 день (тот же лаг);
    #   avg_prev ≡ average + окно = купонный период (окно [start−lag−W, start−lag)
    #              зафиксировано на старте — буквальная семантика avg_prev).
    # Потребители видят только average/month_start + окно. Легаси-значения из
    # старых ручных строк/парсера/калибратора приводятся здесь.
    if out["coupon_mode"] == "point":
        out["coupon_mode"] = "average"
        if out["avg_window_days"] is None:
            out["avg_window_days"] = 1
    elif out["coupon_mode"] == "avg_prev":
        w = out["avg_window_days"]
        if not w:
            try:
                from services import instruments_registry as _reg
                w = (_reg.calc_params_map().get(isin) or {}).get("coupon_period_days")
            except Exception:
                w = None
        if w:
            out["coupon_mode"] = "average"
            out["avg_window_days"] = int(w)
        # период неизвестен → оставляем avg_prev (движок поддерживает legacy)
    return out


# Типы переменной ставки Cbonds, при которых купон ПОСЛЕ оферты неопределён
# (эмитент выставит заново) → оценка к оферте. Для обычного флоатера
# («Плавающая ставка», формула КС/RUONIA+m до погашения) пут — лишь опция
# ликвидности: резать поток нельзя (НРД оценивает к погашению; резка давала
# ложные отрицательные SM у премиальных бумаг с путом).
_RESET_VAR_TYPES = ("определяется решением эмитента", "изменение ставки при условии")


def cut_at_offer(isin: str) -> bool:
    """Резать ли поток флоатера по оферте: только если формула купона после
    оферты неизвестна (пересмотр эмитентом). Неизвестный var_type → не резать."""
    vt = (params(isin).get("var_type") or "").lower()
    if params(isin).get("cut_at_offer") is not None:   # ручной оверрайд
        return bool(params(isin)["cut_at_offer"])
    return any(k in vt for k in _RESET_VAR_TYPES)


def reload():
    """Сброс кэшей (после обновления файлов)."""
    global _cbonds_cache, _manual_cache
    _cbonds_cache = None
    _manual_cache = None
